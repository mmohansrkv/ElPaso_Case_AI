import os
import re
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime

import pandas as pd
from bs4 import BeautifulSoup, NavigableString, Tag
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

# =========================
# COLORS & THEME
# =========================
BG_MAIN      = "#F7F8FA"
BG_CARD      = "#FFFFFF"
BG_SIDEBAR   = "#1D3A6A"
BG_LOG       = "#1E2330"
ACCENT       = "#185FA5"
ACCENT_HOVER = "#1D3A6A"
GREEN        = "#3B6D11"
GREEN_BG     = "#EAF3DE"
RED          = "#A32D2D"
RED_BG       = "#FCEBEB"
AMBER        = "#854F0B"
AMBER_BG     = "#FAEEDA"
PURPLE       = "#5B21B6"
PURPLE_BG    = "#EDE9FE"
ORANGE       = "#92400E"
ORANGE_BG    = "#FEF3C7"
TEXT_PRI     = "#1A1A1A"
TEXT_SEC     = "#6B7280"
TEXT_MUTED   = "#9CA3AF"
BORDER       = "#E5E7EB"

LOG_TEXT     = "#D1FAE5"
LOG_WARN     = "#FDE68A"
LOG_ERR      = "#FCA5A5"
LOG_INFO     = "#93C5FD"

FIELDS = [
    "CaseNumber", "Citation", "CourtCaseType", "CaseStatus",
    "NameRaw", "StatuteDescription",
    "Pleadate", "Pleadate2", "PleaNumber", "Plea",
    "PleaNumber2", "Plea2",
    "Dispositiondate", "Dispositiondate2",
    "DispositionNumber", "Disposition",
    "DispositionNumber2", "Disposition2",
    "AmendedDispositiondate", "AmendedDispositiondate2",
    "AmendedDispositionNumber", "AmendedDisposition",
    "AmendedDisposition2number", "AmendedDisposition2",
    "AmendedPleadate", "AmendedPleadate2",
    "AmendedPleaNumber", "AmendedPlea",
    "AmendedPleaNumber2", "AmendedPlea2",
]

# =========================
# HELPERS
# =========================
def clean(text):
    if pd.isna(text):
        return ""
    return re.sub(r"\s+", " ", str(text)).strip()

def normalize(text):
    return clean(text).lower()

def normalize_case(text):
    return re.sub(r"[.\-\s]", "", clean(text)).lower()

def compare(a, b):
    return "MATCH" if normalize(a) == normalize(b) else "MISMATCH"

# =========================
# CHARGE PARSER
# =========================
def parse_charges_from_outer_div(outer_div):
    charges        = []
    current_number = ""
    current_value  = ""
    for child in outer_div.children:
        if isinstance(child, NavigableString):
            text = clean(str(child))
            num_match = re.match(r'^(\d+)\.', text)
            if num_match:
                if current_number:
                    charges.append((current_number, current_value))
                current_number = num_match.group(1)
                current_value  = ""
        elif isinstance(child, Tag):
            style = child.get("style", "")
            if "padding-left: 40px" in style:
                val = clean(child.text)
                if val and not current_value:
                    current_value = val
    if current_number:
        charges.append((current_number, current_value))
    return charges

# =========================
# DISPOSITIONS EXTRACTION
# =========================
def extract_dispositions_from_html(soup):
    final = {k: "" for k in [
        "Pleadate","Pleadate2","Dispositiondate","Dispositiondate2",
        "AmendedDispositiondate","AmendedPleadate","AmendedPleadate2",
        "PleaNumber","Plea","PleaNumber2","Plea2",
        "DispositionNumber","Disposition","DispositionNumber2","Disposition2",
        "AmendedDispositionNumber","AmendedDisposition",
        "AmendedPleaNumber","AmendedPlea","AmendedPleaNumber2","AmendedPlea2",
        "AmendedDispositiondate2","AmendedDisposition2number","AmendedDisposition2",
    ]}
    disp_table = None
    for table in soup.find_all("table"):
        if table.find("th", id="CDisp"):
            disp_table = table
            break
    if not disp_table:
        return final
    for row in disp_table.find_all("tr"):
        th = row.find("th", class_="ssTableHeaderLabel")
        if not th:
            continue
        date = clean(th.text)
        if not re.match(r'\d{2}/\d{2}/\d{4}', date):
            continue
        tds = row.find_all("td")
        if not tds:
            continue
        content_td = tds[-1]
        bold = content_td.find("b")
        if not bold:
            continue
        type_text = clean(bold.text)
        if "Amended Disposition" in type_text:
            etype = "AmendedDisposition"
        elif "Amended Plea" in type_text:
            etype = "AmendedPlea"
        elif "Disposition" in type_text:
            etype = "Disposition"
        elif "Plea" in type_text:
            etype = "Plea"
        else:
            continue
        all_charges = []
        outer_divs = content_td.find_all(
            "div", style=lambda s: s and "padding-left: 10px" in s
        )
        for outer_div in outer_divs:
            all_charges.extend(parse_charges_from_outer_div(outer_div))
        for number, value in all_charges:
            if etype == "Disposition":
                if not final["Dispositiondate"]:
                    final["Dispositiondate"]   = date
                    final["DispositionNumber"] = number
                    final["Disposition"]       = value
                elif not final["Dispositiondate2"]:
                    final["Dispositiondate2"]   = date
                    final["DispositionNumber2"] = number
                    final["Disposition2"]       = value
            elif etype == "Plea":
                if not final["Pleadate"]:
                    final["Pleadate"]   = date
                    final["PleaNumber"] = number
                    final["Plea"]       = value
                elif not final["Pleadate2"]:
                    final["Pleadate2"]   = date
                    final["PleaNumber2"] = number
                    final["Plea2"]       = value
            elif etype == "AmendedDisposition":
                if not final["AmendedDispositiondate"]:
                    final["AmendedDispositiondate"]   = date
                    final["AmendedDispositionNumber"] = number
                    final["AmendedDisposition"]       = value
                elif not final["AmendedDispositiondate2"]:
                    final["AmendedDispositiondate2"]   = date
                    final["AmendedDisposition2number"] = number
                    final["AmendedDisposition2"]       = value
            elif etype == "AmendedPlea":
                if not final["AmendedPleadate"]:
                    final["AmendedPleadate"]   = date
                    final["AmendedPleaNumber"] = number
                    final["AmendedPlea"]       = value
                elif not final["AmendedPleadate2"]:
                    final["AmendedPleadate2"]   = date
                    final["AmendedPleaNumber2"] = number
                    final["AmendedPlea2"]       = value
    return final

# =========================
# LOAD HTML
# =========================
def load_html(folder, log_fn):
    records = []
    all_files    = os.listdir(folder)
    main_files   = [f for f in all_files if f.startswith("list_")   and f.endswith(".html")]
    detail_files = {f for f in all_files if f.startswith("detail_") and f.endswith(".html")}
    log_fn(f"  📄 Main HTML files  : {len(main_files)}", "info")
    log_fn(f"  📄 Detail files     : {len(detail_files)}", "info")
    for file in main_files:
        case_part = file[len("list_"):].replace(".html", "")
        # FIXED: Removed the backslash before the closing bracket in the regex pattern
        detail_fn = next(
            (f for f in detail_files
             if re.match(r'^detail_' + re.escape(case_part) + r'[_.]', f)),
            None
        )
        dispositions = {k: "" for k in [
            "Pleadate","Pleadate2","Dispositiondate","Dispositiondate2",
            "AmendedDispositiondate","AmendedPleadate","AmendedPleadate2",
            "PleaNumber","Plea","PleaNumber2","Plea2",
            "DispositionNumber","Disposition","DispositionNumber2","Disposition2",
            "AmendedDispositionNumber","AmendedDisposition",
            "AmendedPleaNumber","AmendedPlea","AmendedPleaNumber2","AmendedPlea2",
            "AmendedDispositiondate2","AmendedDisposition2number","AmendedDisposition2",
        ]}
        if detail_fn and detail_fn in detail_files:
            detail_path = os.path.join(folder, detail_fn)
            with open(detail_path, "r", encoding="utf-8", errors="ignore") as f:
                detail_soup = BeautifulSoup(f, "html.parser")
            dispositions = extract_dispositions_from_html(detail_soup)
            log_fn(f"  ✅ Dispositions: {detail_fn}", "ok")
        else:
            log_fn(f"  ⚠️  No detail file for: {file}", "warn")
        main_path = os.path.join(folder, file)
        with open(main_path, "r", encoding="utf-8", errors="ignore") as f:
            main_soup = BeautifulSoup(f, "html.parser")
        for row in main_soup.find_all("tr"):
            cols = row.find_all("td")
            if len(cols) >= 6:
                try:
                    case_no = clean(cols[0].text)
                    if not re.match(r'^[\d.]+[-]\d+-[A-Z]+$', case_no):
                        continue
                    citation     = clean(cols[1].text)
                    name         = clean(cols[2].text)
                    filed_divs   = cols[3].find_all("div")
                    filing_date  = clean(filed_divs[0].text) if len(filed_divs) > 0 else ""
                    jurisdiction = clean(filed_divs[1].text) if len(filed_divs) > 1 else ""
                    type_divs    = cols[4].find_all("div")
                    case_type    = clean(type_divs[0].text) if len(type_divs) > 0 else ""
                    status       = clean(type_divs[1].text) if len(type_divs) > 1 else ""
                    charge       = clean(cols[5].text)
                    base_rec = {
                        "SourceHTML": file, "SourceDetail": detail_fn or "",
                        "CaseNumber": case_no, "CaseNumber_norm": normalize_case(case_no),
                        "Citation": citation, "FilingDate": filing_date,
                        "Jurisdiction": jurisdiction, "CourtCaseType": case_type,
                        "CaseStatus": status, "NameRaw": name, "StatuteDescription": charge,
                    }
                    records.append({**base_rec, **dispositions})
                except:
                    pass
    return pd.DataFrame(records)

# =========================
# LOAD EXCEL
# =========================
def load_excel(folder, log_fn):
    all_data = []
    for file in os.listdir(folder):
        if file.endswith(".xlsx"):
            path = os.path.join(folder, file)
            try:
                df = pd.read_excel(path, dtype=str).fillna("")
                df["SourceExcel"] = file
                all_data.append(df)
            except Exception as e:
                log_fn(f"  ⚠️ Error reading {file}: {e}", "warn")
    if not all_data:
        return pd.DataFrame()
    df = pd.concat(all_data, ignore_index=True)
    if "CaseNumber" in df.columns:
        df["CaseNumber_norm"] = df["CaseNumber"].apply(normalize_case)
    return df

# =========================
# LOAD CLIENT INPUT FILE
# =========================
def load_client(filepath, log_fn):
    if not filepath or not os.path.isfile(filepath):
        log_fn("  ⚠️  No Client Input file provided — ClientMatch column will be skipped.", "warn")
        return set(), pd.DataFrame(), {}
    try:
        df = pd.read_excel(filepath, dtype=str).fillna("")
        if "CaseNumber" in df.columns:
            col = "CaseNumber"
        else:
            col = df.columns[0]
            log_fn(f"  ⚠️  'CaseNumber' column not found in client file; using first column: '{col}'", "warn")
        norm_to_original = {}
        for val in df[col]:
            raw = clean(val)
            norm = normalize_case(raw)
            if norm and norm not in norm_to_original:
                norm_to_original[norm] = raw
        case_set = set(norm_to_original.keys())
        case_set.discard("")
        log_fn(f"  ✅ Client input loaded: {len(case_set)} unique case numbers from '{os.path.basename(filepath)}'", "ok")
        return case_set, df, norm_to_original
    except Exception as e:
        log_fn(f"  ❌ Error reading client file: {e}", "err")
        return set(), pd.DataFrame(), {}

# =========================
# VALIDATION
# =========================
def validate(html_df, excel_df, active_fields, client_cases: set, norm_to_original: dict = None):
    results = []
    compare_fields = [f for f in active_fields if f in html_df.columns]
    use_client = len(client_cases) > 0
    if norm_to_original is None:
        norm_to_original = {}
    for _, h in html_df.iterrows():
        match = excel_df[excel_df["CaseNumber_norm"] == h["CaseNumber_norm"]]
        client_match = ""
        client_casenumber_raw = ""
        if use_client:
            client_match = "FOUND" if h["CaseNumber_norm"] in client_cases else "NOT IN CLIENT"
            client_casenumber_raw = norm_to_original.get(h["CaseNumber_norm"], "")
        if match.empty:
            row = {
                "SourceHTML":             h["SourceHTML"],
                "CaseNumber":             h["CaseNumber"],
                "ClientInputCaseNumber":  client_casenumber_raw,
                "ValidationStatus":       "NOT FOUND",
            }
            if use_client:
                row["ClientCaseNumber_Result"] = compare(client_casenumber_raw, h["CaseNumber"]) if client_casenumber_raw else "—"
                row["ClientMatch"] = client_match
            results.append(row)
            continue
        e = match.iloc[0]
        html_filed  = clean(str(h.get("FilingDate","")) + " " + str(h.get("Jurisdiction","")))
        excel_filed = clean(str(e.get("FilingDate","")) + " " + str(e.get("Jurisdiction","")))
        row = {
            "SourceHTML":            h["SourceHTML"],
            "SourceExcel":           e["SourceExcel"],
            "CaseNumber":            h["CaseNumber"],
            "ClientInputCaseNumber": client_casenumber_raw,
            "Filed_Result":          compare(html_filed, excel_filed),
        }
        for field in compare_fields:
            row[f"{field}_Result"] = compare(h.get(field,""), e.get(field,""))
        checks = [v for k, v in row.items() if k.endswith("_Result")]
        row["ValidationStatus"] = "MATCH" if all(v == "MATCH" for v in checks) else "MISMATCH"
        if use_client:
            output_case = clean(str(e.get("CaseNumber", "")))
            row["ClientCaseNumber_Result"] = compare(client_casenumber_raw, output_case) if client_casenumber_raw else "—"
            row["ClientMatch"] = client_match
        results.append(row)
    result_df = pd.DataFrame(results)
    if use_client and not result_df.empty:
        found_norms = set(result_df["CaseNumber"].apply(normalize_case))
        missing_rows = []
        for norm in client_cases:
            if norm not in found_norms:
                original = norm_to_original.get(norm, norm) if norm_to_original else norm
                missing_rows.append({
                    "SourceHTML":             "—",
                    "CaseNumber":             norm,
                    "ClientInputCaseNumber":  original,
                    "ClientCaseNumber_Result":"—",
                    "ValidationStatus":       "No cases matched your search criteria",
                    "ClientMatch":            "FOUND",
                })
        if missing_rows:
            missing_df = pd.DataFrame(missing_rows)
            result_df  = pd.concat([result_df, missing_df], ignore_index=True)
    return result_df

# =========================
# BUILD ERRORS SHEET
# =========================
def build_errors_df(result_df, html_df, excel_df):
    error_rows = []
    html_lookup  = {r["CaseNumber_norm"]: r for _, r in html_df.iterrows()}
    excel_lookup = {r["CaseNumber_norm"]: r for _, r in excel_df.iterrows()}
    mismatch_result_cols = [c for c in result_df.columns if c.endswith("_Result")]
    for _, row in result_df.iterrows():
        status = row.get("ValidationStatus", "")
        if status != "MISMATCH":
            continue
        case_number = row.get("CaseNumber", "")
        case_norm   = normalize_case(case_number)
        html_row  = html_lookup.get(case_norm, {})
        excel_row = excel_lookup.get(case_norm, {})
        for res_col in mismatch_result_cols:
            if row.get(res_col, "") != "MISMATCH":
                continue
            field = res_col[:-len("_Result")]
            if field == "Filed":
                html_val  = clean(str(html_row.get("FilingDate", "")) + " " + str(html_row.get("Jurisdiction", "")))
                excel_val = clean(str(excel_row.get("FilingDate", "")) + " " + str(excel_row.get("Jurisdiction", "")))
            else:
                html_val  = clean(str(html_row.get(field, "")))
                excel_val = clean(str(excel_row.get(field, "")))
            error_rows.append({
                "CaseNumber": case_number,
                "Field":      field,
                "HTML":       html_val,
                "Excel":      excel_val,
            })
    if not error_rows:
        return pd.DataFrame(columns=["CaseNumber", "Field", "HTML", "Excel"])
    return pd.DataFrame(error_rows, columns=["CaseNumber", "Field", "HTML", "Excel"])

# =========================
# CLIENT INPUT vs EXCEL MATCH SHEET
# =========================
def build_client_excel_match_df(client_df, excel_df):
    if client_df is None or client_df.empty:
        return pd.DataFrame(columns=["ClientInputCaseNumber", "ExcelCaseNumber", "MatchStatus", "SourceExcel"])
    col = "CaseNumber" if "CaseNumber" in client_df.columns else client_df.columns[0]
    excel_lookup = {}
    for _, erow in excel_df.iterrows():
        norm = normalize_case(str(erow.get("CaseNumber", "")))
        if norm and norm not in excel_lookup:
            excel_lookup[norm] = erow
    rows = []
    for val in client_df[col]:
        raw = clean(str(val))
        if not raw:
            continue
        norm = normalize_case(raw)
        erow = excel_lookup.get(norm)
        if erow is not None:
            rows.append({
                "ClientInputCaseNumber": raw,
                "ExcelCaseNumber":       clean(str(erow.get("CaseNumber", ""))),
                "MatchStatus":           "MATCH",
                "SourceExcel":           clean(str(erow.get("SourceExcel", ""))),
            })
        else:
            rows.append({
                "ClientInputCaseNumber": raw,
                "ExcelCaseNumber":       "",
                "MatchStatus":           "MISMATCH",
                "SourceExcel":           "",
            })
    return pd.DataFrame(rows, columns=["ClientInputCaseNumber", "ExcelCaseNumber", "MatchStatus", "SourceExcel"])

# =========================
# BUILD DATA FILL RATE COMPARISON SHEET
# =========================
def build_fill_rate_df(excel_df, active_fields):
    TARGET_RANGES = {
        "CaseNumber": (100, 100),
        "Citation": (90, 100),
        "FilingDate": (90, 100),
        "Jurisdiction": (90, 100),
        "CourtCaseType": (90, 100),
        "CaseStatus": (90, 100),
        "NameRaw": (90, 100),
        "StatuteDescription": (90, 100),
        "Pleadate": (25, 35),
        "Dispositiondate": (25, 35),
        "PleaNumber": (25, 35),
        "Plea": (25, 35),
        "DispositionNumber": (25, 35),
        "Disposition": (25, 35)
    }
    total_excel = len(excel_df)
    common_fields = [f for f in active_fields if f in excel_df.columns]
    rows = []
    for field in common_fields:
        excel_filled = excel_df[field].apply(
            lambda x: pd.notna(x) and str(x).strip() != ""
        ).sum()
        excel_pct = round((excel_filled / total_excel * 100), 2) if total_excel > 0 else 0
        t_min, t_max = TARGET_RANGES.get(field, (0, 10))
        if t_min == t_max:
            target_str = f"{t_min}%"
        else:
            target_str = f"{t_min}-{t_max}%"
        if excel_pct < t_min:
            status = "Below Range"
        elif excel_pct > t_max:
            status = "Above Range"
        else:
            status = "Within Range"
        rows.append({
            "Field": field,
            "Excel Fill Count": excel_filled,
            "Excel Fill %": excel_pct,
            "Expected Target Range": target_str,
            "Status": status,
        })
    if not rows:
        return pd.DataFrame(columns=["Field", "Excel Fill Count", "Excel Fill %", "Expected Target Range", "Status"])
    return pd.DataFrame(rows)

# =========================
# SAVE OUTPUT
# =========================
def save_output(output_file, html_df, excel_df, result_df, client_df, active_fields):
    errors_df = build_errors_df(result_df, html_df, excel_df)
    client_excel_df = build_client_excel_match_df(client_df, excel_df)
    fill_rate_df = build_fill_rate_df(excel_df, active_fields)
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        html_df.to_excel(writer,   sheet_name="HTML",   index=False)
        excel_df.to_excel(writer,  sheet_name="Excel",  index=False)
        result_df.to_excel(writer, sheet_name="Result", index=False)
        errors_df.to_excel(writer, sheet_name="Observation", index=False)
        fill_rate_df.to_excel(writer, sheet_name="FillRate_Comparison", index=False)
        if client_df is not None and not client_df.empty:
            client_df.to_excel(writer, sheet_name="ClientInput", index=False)
        if not client_excel_df.empty:
            client_excel_df.to_excel(writer, sheet_name="ClientInput_vs_Excel", index=False)
    
    wb = load_workbook(output_file)
    
    # ── Style FillRate_Comparison sheet ──
    if "FillRate_Comparison" in wb.sheetnames:
        ws_fr = wb["FillRate_Comparison"]
        header_fill = PatternFill(start_color="1D3A6A", end_color="1D3A6A", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        thin_border = Border(
            left=Side(style='thin', color='D1D5DB'),
            right=Side(style='thin', color='D1D5DB'),
            top=Side(style='thin', color='D1D5DB'),
            bottom=Side(style='thin', color='D1D5DB')
        )
        status_col_idx = None
        pct_col_idx = None
        for idx, cell in enumerate(ws_fr[1], 1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            if cell.value == "Status":
                status_col_idx = idx
            elif cell.value == "Excel Fill %":
                pct_col_idx = idx
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        for row in ws_fr.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            if status_col_idx:
                status_cell = row[status_col_idx - 1]
                pct_cell = row[pct_col_idx - 1] if pct_col_idx else None
                if status_cell.value == "Within Range":
                    status_cell.fill = green_fill
                    if pct_cell: pct_cell.fill = green_fill
                elif status_cell.value == "Below Range":
                    status_cell.fill = red_fill
                    if pct_cell: pct_cell.fill = red_fill
                elif status_cell.value == "Above Range":
                    status_cell.fill = yellow_fill
                    if pct_cell: pct_cell.fill = yellow_fill
        for col in ws_fr.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 4, 30)
            ws_fr.column_dimensions[column_letter].width = adjusted_width
        ws_fr.freeze_panes = "A2"
        ws_fr.auto_filter.ref = ws_fr.dimensions

    # ── Style all other sheets ──
    for sheet_name in wb.sheetnames:
        if sheet_name == "FillRate_Comparison": continue
        ws = wb[sheet_name]
        if ws.max_row is None or ws.max_row < 1 or ws.max_column is None or ws.max_column < 1: continue
        for col in ws.columns:
            max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)
        for cell in ws[1]: cell.font = Font(bold=True)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        if ws.max_row >= 1 and ws.max_column >= 1:
            table_range = f"A1:{ws.cell(row=1, column=ws.max_column).column_letter}{ws.max_row}"
            safe_name = re.sub(r"[^A-Za-z0-9_]", "_", sheet_name) + "Table"
            if not safe_name[0].isalpha() and safe_name[0] != "_": safe_name = "_" + safe_name
            tbl = Table(displayName=safe_name, ref=table_range)
            tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
            ws.add_table(tbl)
    
    # Colour Result sheet
    ws = wb["Result"]
    green_fill  = PatternFill(start_color="C6EFCE", fill_type="solid")
    red_fill    = PatternFill(start_color="FFC7CE", fill_type="solid")
    orange_fill = PatternFill(start_color="FFD700", fill_type="solid")
    grey_fill   = PatternFill(start_color="E5E7EB", fill_type="solid")
    client_col_idx = status_col_idx = client_case_res_idx = None
    for cell in ws[1]:
        if cell.value == "ClientMatch": client_col_idx = cell.column
        if cell.value == "ValidationStatus": status_col_idx = cell.column
        if cell.value == "ClientCaseNumber_Result": client_case_res_idx = cell.column
    for row in ws.iter_rows(min_row=2):
        status_val = row[status_col_idx - 1].value if status_col_idx else ""
        for cell in row:
            if cell.value == "MATCH": cell.fill = green_fill
            elif cell.value == "MISMATCH": cell.fill = red_fill
            elif status_val == "No cases matched your search criteria": cell.fill = orange_fill
            elif cell.column == client_col_idx:
                if cell.value == "FOUND": cell.fill = green_fill
                elif cell.value == "NOT IN CLIENT": cell.fill = grey_fill
            if client_case_res_idx and cell.column == client_case_res_idx:
                if cell.value == "MATCH": cell.fill = green_fill
                elif cell.value == "MISMATCH": cell.fill = red_fill

    # ── Colour Errors sheet ──────────────────────────────────────────────────
    if "Errors" in wb.sheetnames:
        ws_err = wb["Errors"]
        html_col_idx = excel_col_idx = None
        for cell in ws_err[1]:
            if cell.value == "HTML": html_col_idx = cell.column
            elif cell.value == "Excel": excel_col_idx = cell.column
        for row in ws_err.iter_rows(min_row=2):
            for cell in row:
                if cell.column == html_col_idx and cell.value not in (None, ""):
                    cell.fill = PatternFill(start_color="DDEBF7", fill_type="solid")
                elif cell.column == excel_col_idx and cell.value not in (None, ""):
                    cell.fill = PatternFill(start_color="FFC7CE", fill_type="solid")

    # ── Colour ClientInput_vs_Excel sheet ────────────────────────────────────
    if "ClientInput_vs_Excel" in wb.sheetnames:
        ws_ce = wb["ClientInput_vs_Excel"]
        match_col_idx = None
        for cell in ws_ce[1]:
            if cell.value == "MatchStatus": match_col_idx = cell.column
        for row in ws_ce.iter_rows(min_row=2):
            status_val = row[match_col_idx - 1].value if match_col_idx else ""
            row_fill = green_fill if status_val == "MATCH" else red_fill
            for cell in row: cell.fill = row_fill
    wb.save(output_file)

# =========================
# GUI APPLICATION
# =========================
class ElPasoApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("El Paso Case Validator")
        self.geometry("1100x720")
        self.minsize(900, 600)
        self.configure(bg=BG_MAIN)
        self.resizable(True, True)
        self.html_path    = tk.StringVar(value=r"D:\Mohan\aug 21\TXElPaso_NewCases_20260417_CaseProfile")
        self.excel_path   = tk.StringVar(value=r"D:\Mohan\aug 21\Output")
        self.output_path  = tk.StringVar(value=r"D:\Mohan\aug 21\Result")
        self.client_path  = tk.StringVar(value="")
        self.status_var   = tk.StringVar(value="Idle")
        self.field_vars   = {f: tk.BooleanVar(value=True) for f in FIELDS}
        self.html_count        = tk.StringVar(value="—")
        self.excel_count       = tk.StringVar(value="—")
        self.match_count       = tk.StringVar(value="—")
        self.mismatch_count    = tk.StringVar(value="—")
        self.notfound_out_count = tk.StringVar(value="—")
        self._result_df = None
        self._filter    = "all"
        self._build_ui()

    def _build_ui(self):
        self._style()
        outer = tk.Frame(self, bg=BG_MAIN)
        outer.pack(fill="both", expand=True)
        sidebar = tk.Frame(outer, bg=BG_SIDEBAR, width=230)
        sidebar.pack(side="left", fill="y")
        sidebar.pack_propagate(False)
        self._build_sidebar(sidebar)
        main = tk.Frame(outer, bg=BG_MAIN)
        main.pack(side="left", fill="both", expand=True)
        self._build_main(main)

    def _style(self):
        s = ttk.Style(self)
        s.theme_use("clam")
        s.configure("Card.TFrame",      background=BG_CARD,   relief="flat")
        s.configure("TLabel",           background=BG_MAIN,   foreground=TEXT_PRI,  font=("Segoe UI", 10))
        s.configure("Muted.TLabel",     background=BG_MAIN,   foreground=TEXT_SEC,  font=("Segoe UI", 9))
        s.configure("Card.TLabel",      background=BG_CARD,   foreground=TEXT_PRI,  font=("Segoe UI", 10))
        s.configure("CardMuted.TLabel", background=BG_CARD,   foreground=TEXT_SEC,  font=("Segoe UI", 9))
        s.configure("Metric.TLabel",    background="#F0F4FA",  foreground=TEXT_PRI,  font=("Segoe UI", 18, "bold"))
        s.configure("MetricLbl.TLabel", background="#F0F4FA",  foreground=TEXT_SEC,  font=("Segoe UI", 9))
        s.configure("Run.TButton",      font=("Segoe UI", 11, "bold"), background=ACCENT, foreground="#FFFFFF", borderwidth=0)
        s.map("Run.TButton", background=[("active", ACCENT_HOVER)])
        s.configure("Treeview",         font=("Segoe UI", 9), rowheight=26, background=BG_CARD, fieldbackground=BG_CARD, foreground=TEXT_PRI)
        s.configure("Treeview.Heading", font=("Segoe UI", 9, "bold"), background="#F0F4FA", foreground=TEXT_SEC, relief="flat")
        s.map("Treeview", background=[("selected", "#DBEAFE")])
        s.configure("green.Horizontal.TProgressbar", troughcolor=BORDER, background=ACCENT, thickness=4)

    def _build_sidebar(self, parent):
        hdr = tk.Frame(parent, bg=BG_SIDEBAR, pady=20, padx=16)
        hdr.pack(fill="x")
        tk.Label(hdr, text="⚖", bg=BG_SIDEBAR, fg="#93C5FD", font=("Segoe UI", 24)).pack(anchor="w")
        tk.Label(hdr, text="El Paso\nCase Validator", bg=BG_SIDEBAR, fg="#FFFFFF", font=("Segoe UI", 13, "bold"), justify="left").pack(anchor="w", pady=(4,0))
        tk.Label(hdr, text="HTML → Excel comparison", bg=BG_SIDEBAR, fg="#93C5FD", font=("Segoe UI", 8)).pack(anchor="w")
        tk.Frame(parent, bg="#2D5499", height=1).pack(fill="x", padx=16, pady=8)
        sf = tk.Frame(parent, bg=BG_SIDEBAR, padx=16)
        sf.pack(fill="x", pady=(0,12))
        tk.Label(sf, text="STATUS", bg=BG_SIDEBAR, fg="#64A0D4", font=("Segoe UI", 8, "bold")).pack(anchor="w")
        self._status_lbl = tk.Label(sf, textvariable=self.status_var, bg=BG_SIDEBAR, fg="#93C5FD", font=("Segoe UI", 10, "bold"))
        self._status_lbl.pack(anchor="w")
        tk.Frame(parent, bg="#2D5499", height=1).pack(fill="x", padx=16, pady=8)
        sel_all_fr = tk.Frame(parent, bg=BG_SIDEBAR, padx=14, pady=4)
        sel_all_fr.pack(fill="x")
        tk.Button(sel_all_fr, text="📁  Select All Paths", command=lambda: self._browse("all"), bg="#2563EB", fg="#FFFFFF", font=("Segoe UI", 9, "bold"), relief="flat", pady=6, cursor="hand2", activebackground="#1D4ED8", activeforeground="#FFFFFF").pack(fill="x")
        tk.Label(parent, text="PATHS", bg=BG_SIDEBAR, fg="#64A0D4", font=("Segoe UI", 8, "bold"), padx=16).pack(anchor="w")
        self._path_row(parent, "HTML Folder", self.html_path, "html")
        self._path_row(parent, "Excel Folder", self.excel_path, "excel")
        self._path_row(parent, "Result File", self.output_path, "output")
        tk.Frame(parent, bg="#2D5499", height=1).pack(fill="x", padx=16, pady=(10, 4))
        tk.Label(parent, text="CLIENT INPUT", bg=BG_SIDEBAR, fg="#64A0D4", font=("Segoe UI", 8, "bold"), padx=16).pack(anchor="w")
        self._path_row(parent, "Client Input File", self.client_path, "client")
        tk.Frame(parent, bg="#2D5499", height=1).pack(fill="x", padx=16, pady=10)
        tk.Label(parent, text="VALIDATION FIELDS", bg=BG_SIDEBAR, fg="#64A0D4", font=("Segoe UI", 8, "bold"), padx=16).pack(anchor="w")
        ff = tk.Frame(parent, bg=BG_SIDEBAR, padx=14)
        ff.pack(fill="both", expand=True)
        canvas = tk.Canvas(ff, bg=BG_SIDEBAR, highlightthickness=0)
        scroll = tk.Scrollbar(ff, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scroll.set)
        scroll.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        inner = tk.Frame(canvas, bg=BG_SIDEBAR)
        canvas.create_window((0,0), window=inner, anchor="nw")
        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        for f in FIELDS:
            cb = tk.Checkbutton(inner, text=f, variable=self.field_vars[f], bg=BG_SIDEBAR, fg="#CBD5E1", selectcolor="#2D5499", activebackground=BG_SIDEBAR, activeforeground="#FFFFFF", font=("Segoe UI", 8), anchor="w", width=22)
            cb.pack(anchor="w", pady=1)

    def _path_row(self, parent, label, var, key):
        fr = tk.Frame(parent, bg=BG_SIDEBAR, padx=14, pady=3)
        fr.pack(fill="x")
        tk.Label(fr, text=label, bg=BG_SIDEBAR, fg="#94A3B8", font=("Segoe UI", 8)).pack(anchor="w")
        row = tk.Frame(fr, bg=BG_SIDEBAR)
        row.pack(fill="x")
        tk.Entry(row, textvariable=var, bg="#253F70", fg="#E2E8F0", insertbackground="#E2E8F0", font=("Segoe UI", 8), relief="flat", bd=4, width=20).pack(side="left", fill="x", expand=True)
        tk.Button(row, text="…", bg="#2D5499", fg="#FFFFFF", font=("Segoe UI", 8, "bold"), relief="flat", padx=6, command=lambda k=key: self._browse(k)).pack(side="left", padx=(3,0))

    def _build_main(self, parent):
        mf = tk.Frame(parent, bg=BG_MAIN, padx=20, pady=16)
        mf.pack(fill="x")
        metrics = [("HTML Rows", self.html_count, "#185FA5"), ("Excel Rows", self.excel_count, "#185FA5"), ("Matched", self.match_count, GREEN), ("Mismatched", self.mismatch_count, RED), ("Not Found in Output", self.notfound_out_count, ORANGE)]
        for i, (lbl, var, color) in enumerate(metrics):
            card = tk.Frame(mf, bg="#F0F4FA", bd=0, relief="flat")
            card.grid(row=0, column=i, padx=6, sticky="nsew")
            mf.columnconfigure(i, weight=1)
            tk.Label(card, text=lbl, bg="#F0F4FA", fg=TEXT_SEC, font=("Segoe UI", 9), pady=8).pack(anchor="w", padx=12)
            tk.Label(card, textvariable=var, bg="#F0F4FA", fg=color, font=("Segoe UI", 20, "bold"), pady=4).pack(anchor="w", padx=12)
        rb = tk.Frame(parent, bg=BG_MAIN, padx=20)
        rb.pack(fill="x")
        self._run_btn = tk.Button(rb, text="▶  Run Validation", command=self._run, bg=ACCENT, fg="#FFFFFF", font=("Segoe UI", 11, "bold"), relief="flat", pady=10, cursor="hand2", activebackground=ACCENT_HOVER, activeforeground="#FFFFFF")
        self._run_btn.pack(fill="x", pady=(0,8))
        self._progress = ttk.Progressbar(rb, style="green.Horizontal.TProgressbar", mode="determinate", maximum=100)
        self._progress.pack(fill="x")
        nb_frame = tk.Frame(parent, bg=BG_MAIN, padx=20, pady=12)
        nb_frame.pack(fill="both", expand=True)
        nb = ttk.Notebook(nb_frame)
        nb.pack(fill="both", expand=True)
        log_tab = tk.Frame(nb, bg=BG_LOG)
        nb.add(log_tab, text="  Run Log  ")
        self._log = tk.Text(log_tab, bg=BG_LOG, fg="#D1FAE5", font=("Consolas", 9), relief="flat", wrap="none", state="disabled", insertbackground="#D1FAE5", bd=8)
        log_scroll = tk.Scrollbar(log_tab, command=self._log.yview)
        self._log.configure(yscrollcommand=log_scroll.set)
        log_scroll.pack(side="right", fill="y")
        self._log.pack(fill="both", expand=True)
        self._log.tag_config("ok",   foreground=LOG_TEXT)
        self._log.tag_config("warn", foreground=LOG_WARN)
        self._log.tag_config("err",  foreground=LOG_ERR)
        self._log.tag_config("info", foreground=LOG_INFO)
        res_tab = tk.Frame(nb, bg=BG_CARD)
        nb.add(res_tab, text="  Results  ")
        self._build_results(res_tab)

    def _build_results(self, parent):
        fbar = tk.Frame(parent, bg=BG_CARD, pady=8, padx=8)
        fbar.pack(fill="x")
        self._filter_btns = {}
        for label in ["All", "MATCH", "MISMATCH", "NOT FOUND", "No cases matched your search criteria"]:
            key = "all" if label == "All" else label
            btn = tk.Button(fbar, text=label, font=("Segoe UI", 9), bg=ACCENT if label=="All" else BORDER, fg="#FFFFFF" if label=="All" else TEXT_SEC, relief="flat", padx=10, pady=4, command=lambda k=key: self._set_filter(k))
            btn.pack(side="left", padx=2)
            self._filter_btns[key] = btn
        tk.Label(fbar, text="Search:", bg=BG_CARD, fg=TEXT_SEC, font=("Segoe UI", 9)).pack(side="left", padx=(10,4))
        self._search_var = tk.StringVar()
        self._search_var.trace_add("write", lambda *_: self._render_table())
        tk.Entry(fbar, textvariable=self._search_var, font=("Segoe UI", 9), bg="#F9FAFB", fg=TEXT_PRI, relief="flat", bd=4, width=22).pack(side="left")
        cols    = ("SourceHTML", "CaseNumber", "ClientInputCaseNumber", "ClientCaseMatch", "SourceExcel", "Status", "ClientMatch", "Issues")
        self._tree = ttk.Treeview(parent, columns=cols, show="headings", selectmode="browse")
        for col in cols: self._tree.heading(col, text=col)
        self._tree.tag_configure("MATCH", background=GREEN_BG, foreground=GREEN)
        self._tree.tag_configure("MISMATCH", background=RED_BG, foreground=RED)
        self._tree.tag_configure("NOT FOUND", background=AMBER_BG, foreground=AMBER)
        self._tree.tag_configure("No cases matched your search criteria", background=ORANGE_BG, foreground=ORANGE)
        self._tree.tag_configure("CLIENT_MISMATCH", background=PURPLE_BG, foreground=PURPLE)
        vsb = ttk.Scrollbar(parent, orient="vertical", command=self._tree.yview)
        hsb = ttk.Scrollbar(parent, orient="horizontal", command=self._tree.xview)
        self._tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        hsb.pack(side="bottom", fill="x")
        vsb.pack(side="right", fill="y")
        self._tree.pack(fill="both", expand=True)

    def _browse(self, key):
        if key == "all": self._browse_all()
        elif key == "output":
            path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files","*.xlsx")])
            if path: self.output_path.set(path)
        elif key == "client":
            path = filedialog.askopenfilename(filetypes=[("Excel files","*.xlsx *.xls")])
            if path: self.client_path.set(path)
        else:
            path = filedialog.askdirectory()
            if path: (self.html_path if key=="html" else self.excel_path).set(path)

    def _browse_all(self):
        base = filedialog.askdirectory(title="Select Root Folder")
        if not base: return
        html_dir = os.path.join(base, "HTML Folder")
        self.html_path.set(html_dir if os.path.isdir(html_dir) else base)
        excel_dir = os.path.join(base, "Excel folder")
        self.excel_path.set(excel_dir if os.path.isdir(excel_dir) else base)
        result_dir = os.path.join(base, "Result")
        os.makedirs(result_dir, exist_ok=True)
        self.output_path.set(result_dir)
        client_dir = os.path.join(base, "Client Input")
        if os.path.isdir(client_dir):
            for fname in sorted(os.listdir(client_dir)):
                if fname.lower().endswith((".xlsx", ".xls")):
                    self.client_path.set(os.path.join(client_dir, fname))
                    break

    def _log_msg(self, msg, tag=""):
        ts = datetime.now().strftime("%H:%M:%S")
        self._log.configure(state="normal")
        self._log.insert("end", f"[{ts}]  {msg}\n", tag)
        self._log.see("end")
        self._log.configure(state="disabled")

    def _set_progress(self, val):
        self._progress["value"] = val
        self.update_idletasks()

    def _set_filter(self, key):
        self._filter = key
        for k, btn in self._filter_btns.items():
            btn.configure(bg=ACCENT if k==key else BORDER, fg="#FFFFFF" if k==key else TEXT_SEC)
        self._render_table()

    def _render_table(self):
        for row in self._tree.get_children(): self._tree.delete(row)
        if self._result_df is None or self._result_df.empty: return
        df = self._result_df
        q = self._search_var.get().lower()
        if self._filter != "all": df = df[df["ValidationStatus"] == self._filter]
        if q:
            mask = df.apply(lambda r: q in str(r.get("CaseNumber","")).lower() or q in str(r.get("SourceHTML","")).lower(), axis=1)
            df = df[mask]
        mismatch_cols = [c for c in df.columns if c.endswith("_Result")]
        for _, row in df.iterrows():
            status = row.get("ValidationStatus", "")
            bad = ", ".join(c.replace("_Result","") for c in mismatch_cols if row.get(c,"") == "MISMATCH") or "—"
            tag = status
            if row.get("ClientCaseNumber_Result", "—") == "MISMATCH" and status == "MATCH": tag = "CLIENT_MISMATCH"
            self._tree.insert("", "end", values=(row.get("SourceHTML", ""), row.get("CaseNumber", ""), row.get("ClientInputCaseNumber", "—"), row.get("ClientCaseNumber_Result", "—"), row.get("SourceExcel", "—"), status, row.get("ClientMatch", "—"), bad), tags=(tag,))

    def _run(self):
        self._run_btn.configure(state="disabled", text="Running…")
        self.status_var.set("Running…")
        self._status_lbl.configure(fg="#FDE68A")
        self._set_progress(0)
        self._log.configure(state="normal")
        self._log.delete("1.0","end")
        self._log.configure(state="disabled")
        threading.Thread(target=self._run_pipeline, daemon=True).start()

    def _run_pipeline(self):
        try:
            html_f = self.html_path.get(); excel_f = self.excel_path.get(); out_f = self.output_path.get(); client_f = self.client_path.get()
            if os.path.isdir(out_f): out_f = os.path.join(out_f, f"Result_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"); self.output_path.set(out_f)
            if not os.path.isdir(html_f) or not os.path.isdir(excel_f): self._log_msg("❌ Check folder paths.", "err"); self._done(error=True); return
            self._log_msg("📂 Loading HTML files…", "info")
            html_df = load_html(html_f, self._log_msg)
            self.html_count.set(str(len(html_df))); self._set_progress(20)
            self._log_msg("📂 Loading Excel files…", "info")
            excel_df = load_excel(excel_f, self._log_msg)
            self.excel_count.set(str(len(excel_df))); self._set_progress(40)
            client_cases, client_df, norm_to_orig = load_client(client_f, self._log_msg); self._set_progress(55)
            if html_df.empty or excel_df.empty: self._log_msg("❌ No data found.", "err"); self._done(error=True); return
            active_f = [f for f, v in self.field_vars.items() if v.get()]
            result_df = validate(html_df, excel_df, active_f, client_cases, norm_to_orig); self._set_progress(80)
            matches = int((result_df["ValidationStatus"]=="MATCH").sum())
            mismatches = int((result_df["ValidationStatus"]=="MISMATCH").sum())
            notfound = int((result_df["ValidationStatus"]=="NOT FOUND").sum())
            notfound_out = int((result_df["ValidationStatus"]=="No cases matched your search criteria").sum())
            self.match_count.set(str(matches)); self.mismatch_count.set(str(mismatches + notfound)); self.notfound_out_count.set(str(notfound_out))
            save_output(out_f, html_df, excel_df, result_df, client_df if not client_df.empty else None, active_f)
            self._log_msg(f"✅ Saved: {out_f}", "ok"); self._set_progress(100); self._result_df = result_df; self.after(0, self._render_table); self._done(error=False)
            if messagebox.askyesno("Done", "Validation complete! Open output file?"): os.startfile(out_f)
        except Exception as ex: self._log_msg(f"❌ Error: {ex}", "err"); self._done(error=True)

    def _done(self, error=False):
        self._run_btn.configure(state="normal", text="▶  Run Validation")
        self.status_var.set("Error" if error else "Complete ✓")
        self._status_lbl.configure(fg=LOG_ERR if error else LOG_TEXT)

if __name__ == "__main__":
    ElPasoApp().mainloop()
