"""
El Paso Case Validator - core logic
Extracted from the original Tkinter desktop app with NO changes to the
parsing / matching / Excel-export logic. Only the GUI layer was removed.
"""
import os
import re

import pandas as pd
from bs4 import BeautifulSoup, NavigableString, Tag
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

# =========================
# FIELDS
# =========================
FIELDS = [
    "CaseNumber", "Citation", "CourtCaseType", "CaseStatus",
    "NameRaw", "StatuteDescription",
    "Pleadate", "Pleadate2", "PleaNumber", "Plea",
    "PleaNumber2", "Plea2",
    "Dispositiondate", "Dispositiondate2",
    "DispositionNumber", "Disposition",
    "DispositionNumber2", "Disposition2",
    "AmendedDispositiondate", "AmendedDispositiondate2",
    "AmendedDispositionNumber", "AmendedDisposition",
    "AmendedDisposition2number", "AmendedDisposition2",
    "AmendedPleadate", "AmendedPleadate2",
    "AmendedPleaNumber", "AmendedPlea",
    "AmendedPleaNumber2", "AmendedPlea2",
]

# =========================
# HELPERS
# =========================
def clean(text):
    if pd.isna(text):
        return ""
    return re.sub(r"\s+", " ", str(text)).strip()

def normalize(text):
    return clean(text).lower()

def normalize_case(text):
    return re.sub(r"[.\-\s]", "", clean(text)).lower()

def compare(a, b):
    return "MATCH" if normalize(a) == normalize(b) else "MISMATCH"

# =========================
# CHARGE PARSER
# =========================
def parse_charges_from_outer_div(outer_div):
    charges        = []
    current_number = ""
    current_value  = ""
    for child in outer_div.children:
        if isinstance(child, NavigableString):
            text = clean(str(child))
            num_match = re.match(r'^(\d+)\.', text)
            if num_match:
                if current_number:
                    charges.append((current_number, current_value))
                current_number = num_match.group(1)
                current_value  = ""
        elif isinstance(child, Tag):
            style = child.get("style", "")
            if "padding-left: 40px" in style:
                val = clean(child.text)
                if val and not current_value:
                    current_value = val
    if current_number:
        charges.append((current_number, current_value))
    return charges

# =========================
# DISPOSITIONS EXTRACTION
# =========================
def extract_dispositions_from_html(soup):
    final = {k: "" for k in [
        "Pleadate","Pleadate2","Dispositiondate","Dispositiondate2",
        "AmendedDispositiondate","AmendedPleadate","AmendedPleadate2",
        "PleaNumber","Plea","PleaNumber2","Plea2",
        "DispositionNumber","Disposition","DispositionNumber2","Disposition2",
        "AmendedDispositionNumber","AmendedDisposition",
        "AmendedPleaNumber","AmendedPlea","AmendedPleaNumber2","AmendedPlea2",
        "AmendedDispositiondate2","AmendedDisposition2number","AmendedDisposition2",
    ]}
    disp_table = None
    for table in soup.find_all("table"):
        if table.find("th", id="CDisp"):
            disp_table = table
            break
    if not disp_table:
        return final
    for row in disp_table.find_all("tr"):
        th = row.find("th", class_="ssTableHeaderLabel")
        if not th:
            continue
        date = clean(th.text)
        if not re.match(r'\d{2}/\d{2}/\d{4}', date):
            continue
        tds = row.find_all("td")
        if not tds:
            continue
        content_td = tds[-1]
        bold = content_td.find("b")
        if not bold:
            continue
        type_text = clean(bold.text)
        if "Amended Disposition" in type_text:
            etype = "AmendedDisposition"
        elif "Amended Plea" in type_text:
            etype = "AmendedPlea"
        elif "Disposition" in type_text:
            etype = "Disposition"
        elif "Plea" in type_text:
            etype = "Plea"
        else:
            continue
        all_charges = []
        outer_divs = content_td.find_all(
            "div", style=lambda s: s and "padding-left: 10px" in s
        )
        for outer_div in outer_divs:
            all_charges.extend(parse_charges_from_outer_div(outer_div))
        for number, value in all_charges:
            if etype == "Disposition":
                if not final["Dispositiondate"]:
                    final["Dispositiondate"]   = date
                    final["DispositionNumber"] = number
                    final["Disposition"]       = value
                elif not final["Dispositiondate2"]:
                    final["Dispositiondate2"]   = date
                    final["DispositionNumber2"] = number
                    final["Disposition2"]       = value
            elif etype == "Plea":
                if not final["Pleadate"]:
                    final["Pleadate"]   = date
                    final["PleaNumber"] = number
                    final["Plea"]       = value
                elif not final["Pleadate2"]:
                    final["Pleadate2"]   = date
                    final["PleaNumber2"] = number
                    final["Plea2"]       = value
            elif etype == "AmendedDisposition":
                if not final["AmendedDispositiondate"]:
                    final["AmendedDispositiondate"]   = date
                    final["AmendedDispositionNumber"] = number
                    final["AmendedDisposition"]       = value
                elif not final["AmendedDispositiondate2"]:
                    final["AmendedDispositiondate2"]   = date
                    final["AmendedDisposition2number"] = number
                    final["AmendedDisposition2"]       = value
            elif etype == "AmendedPlea":
                if not final["AmendedPleadate"]:
                    final["AmendedPleadate"]   = date
                    final["AmendedPleaNumber"] = number
                    final["AmendedPlea"]       = value
                elif not final["AmendedPleadate2"]:
                    final["AmendedPleadate2"]   = date
                    final["AmendedPleaNumber2"] = number
                    final["AmendedPlea2"]       = value
    return final

# =========================
# LOAD HTML
# =========================
def load_html(folder, log_fn):
    records = []
    all_files    = os.listdir(folder)
    main_files   = [f for f in all_files if f.startswith("list_")   and f.endswith(".html")]
    detail_files = {f for f in all_files if f.startswith("detail_") and f.endswith(".html")}
    log_fn(f"  Main HTML files  : {len(main_files)}", "info")
    log_fn(f"  Detail files     : {len(detail_files)}", "info")
    for file in main_files:
        case_part = file[len("list_"):].replace(".html", "")
        detail_fn = next(
            (f for f in detail_files
             if re.match(r'^detail_' + re.escape(case_part) + r'[_.]', f)),
            None
        )
        dispositions = {k: "" for k in [
            "Pleadate","Pleadate2","Dispositiondate","Dispositiondate2",
            "AmendedDispositiondate","AmendedPleadate","AmendedPleadate2",
            "PleaNumber","Plea","PleaNumber2","Plea2",
            "DispositionNumber","Disposition","DispositionNumber2","Disposition2",
            "AmendedDispositionNumber","AmendedDisposition",
            "AmendedPleaNumber","AmendedPlea","AmendedPleaNumber2","AmendedPlea2",
            "AmendedDispositiondate2","AmendedDisposition2number","AmendedDisposition2",
        ]}
        if detail_fn and detail_fn in detail_files:
            detail_path = os.path.join(folder, detail_fn)
            with open(detail_path, "r", encoding="utf-8", errors="ignore") as f:
                detail_soup = BeautifulSoup(f, "html.parser")
            dispositions = extract_dispositions_from_html(detail_soup)
            log_fn(f"  Dispositions: {detail_fn}", "ok")
        else:
            log_fn(f"  No detail file for: {file}", "warn")
        main_path = os.path.join(folder, file)
        with open(main_path, "r", encoding="utf-8", errors="ignore") as f:
            main_soup = BeautifulSoup(f, "html.parser")
        for row in main_soup.find_all("tr"):
            cols = row.find_all("td")
            if len(cols) >= 6:
                try:
                    case_no = clean(cols[0].text)
                    if not re.match(r'^[\d.]+[-]\d+-[A-Z]+$', case_no):
                        continue
                    citation     = clean(cols[1].text)
                    name         = clean(cols[2].text)
                    filed_divs   = cols[3].find_all("div")
                    filing_date  = clean(filed_divs[0].text) if len(filed_divs) > 0 else ""
                    jurisdiction = clean(filed_divs[1].text) if len(filed_divs) > 1 else ""
                    type_divs    = cols[4].find_all("div")
                    case_type    = clean(type_divs[0].text) if len(type_divs) > 0 else ""
                    status       = clean(type_divs[1].text) if len(type_divs) > 1 else ""
                    charge       = clean(cols[5].text)
                    base_rec = {
                        "SourceHTML": file, "SourceDetail": detail_fn or "",
                        "CaseNumber": case_no, "CaseNumber_norm": normalize_case(case_no),
                        "Citation": citation, "FilingDate": filing_date,
                        "Jurisdiction": jurisdiction, "CourtCaseType": case_type,
                        "CaseStatus": status, "NameRaw": name, "StatuteDescription": charge,
                    }
                    records.append({**base_rec, **dispositions})
                except Exception:
                    pass
    return pd.DataFrame(records)

# =========================
# LOAD EXCEL
# =========================
def load_excel(folder, log_fn):
    all_data = []
    for file in os.listdir(folder):
        if file.endswith(".xlsx"):
            path = os.path.join(folder, file)
            try:
                df = pd.read_excel(path, dtype=str).fillna("")
                df["SourceExcel"] = file
                all_data.append(df)
            except Exception as e:
                log_fn(f"  Error reading {file}: {e}", "warn")
    if not all_data:
        return pd.DataFrame()
    df = pd.concat(all_data, ignore_index=True)
    if "CaseNumber" in df.columns:
        df["CaseNumber_norm"] = df["CaseNumber"].apply(normalize_case)
    return df

# =========================
# LOAD CLIENT INPUT FILE
# =========================
def load_client(filepath, log_fn):
    if not filepath or not os.path.isfile(filepath):
        log_fn("  No Client Input file provided - ClientMatch column will be skipped.", "warn")
        return set(), pd.DataFrame(), {}
    try:
        df = pd.read_excel(filepath, dtype=str).fillna("")
        if "CaseNumber" in df.columns:
            col = "CaseNumber"
        else:
            col = df.columns[0]
            log_fn(f"  'CaseNumber' column not found in client file; using first column: '{col}'", "warn")
        norm_to_original = {}
        for val in df[col]:
            raw = clean(val)
            norm = normalize_case(raw)
            if norm and norm not in norm_to_original:
                norm_to_original[norm] = raw
        case_set = set(norm_to_original.keys())
        case_set.discard("")
        log_fn(f"  Client input loaded: {len(case_set)} unique case numbers from '{os.path.basename(filepath)}'", "ok")
        return case_set, df, norm_to_original
    except Exception as e:
        log_fn(f"  Error reading client file: {e}", "err")
        return set(), pd.DataFrame(), {}

# =========================
# VALIDATION
# =========================
def validate(html_df, excel_df, active_fields, client_cases: set, norm_to_original: dict = None):
    results = []
    compare_fields = [f for f in active_fields if f in html_df.columns]
    use_client = len(client_cases) > 0
    if norm_to_original is None:
        norm_to_original = {}
    for _, h in html_df.iterrows():
        match = excel_df[excel_df["CaseNumber_norm"] == h["CaseNumber_norm"]]
        client_match = ""
        client_casenumber_raw = ""
        if use_client:
            client_match = "FOUND" if h["CaseNumber_norm"] in client_cases else "NOT IN CLIENT"
            client_casenumber_raw = norm_to_original.get(h["CaseNumber_norm"], "")
        if match.empty:
            row = {
                "SourceHTML":             h["SourceHTML"],
                "CaseNumber":             h["CaseNumber"],
                "ClientInputCaseNumber":  client_casenumber_raw,
                "ValidationStatus":       "NOT FOUND",
            }
            if use_client:
                row["ClientCaseNumber_Result"] = compare(client_casenumber_raw, h["CaseNumber"]) if client_casenumber_raw else "-"
                row["ClientMatch"] = client_match
            results.append(row)
            continue
        e = match.iloc[0]
        html_filed  = clean(str(h.get("FilingDate","")) + " " + str(h.get("Jurisdiction","")))
        excel_filed = clean(str(e.get("FilingDate","")) + " " + str(e.get("Jurisdiction","")))
        row = {
            "SourceHTML":            h["SourceHTML"],
            "SourceExcel":           e["SourceExcel"],
            "CaseNumber":            h["CaseNumber"],
            "ClientInputCaseNumber": client_casenumber_raw,
            "Filed_Result":          compare(html_filed, excel_filed),
        }
        for field in compare_fields:
            row[f"{field}_Result"] = compare(h.get(field,""), e.get(field,""))
        checks = [v for k, v in row.items() if k.endswith("_Result")]
        row["ValidationStatus"] = "MATCH" if all(v == "MATCH" for v in checks) else "MISMATCH"
        if use_client:
            output_case = clean(str(e.get("CaseNumber", "")))
            row["ClientCaseNumber_Result"] = compare(client_casenumber_raw, output_case) if client_casenumber_raw else "-"
            row["ClientMatch"] = client_match
        results.append(row)
    result_df = pd.DataFrame(results)
    if use_client and not result_df.empty:
        found_norms = set(result_df["CaseNumber"].apply(normalize_case))
        missing_rows = []
        for norm in client_cases:
            if norm not in found_norms:
                original = norm_to_original.get(norm, norm) if norm_to_original else norm
                missing_rows.append({
                    "SourceHTML":             "-",
                    "CaseNumber":             norm,
                    "ClientInputCaseNumber":  original,
                    "ClientCaseNumber_Result":"-",
                    "ValidationStatus":       "No cases matched your search criteria",
                    "ClientMatch":            "FOUND",
                })
        if missing_rows:
            missing_df = pd.DataFrame(missing_rows)
            result_df  = pd.concat([result_df, missing_df], ignore_index=True)
    return result_df

# =========================
# BUILD ERRORS SHEET
# =========================
def build_errors_df(result_df, html_df, excel_df):
    error_rows = []
    html_lookup  = {r["CaseNumber_norm"]: r for _, r in html_df.iterrows()}
    excel_lookup = {r["CaseNumber_norm"]: r for _, r in excel_df.iterrows()}
    mismatch_result_cols = [c for c in result_df.columns if c.endswith("_Result")]
    for _, row in result_df.iterrows():
        status = row.get("ValidationStatus", "")
        if status != "MISMATCH":
            continue
        case_number = row.get("CaseNumber", "")
        case_norm   = normalize_case(case_number)
        html_row  = html_lookup.get(case_norm, {})
        excel_row = excel_lookup.get(case_norm, {})
        for res_col in mismatch_result_cols:
            if row.get(res_col, "") != "MISMATCH":
                continue
            field = res_col[:-len("_Result")]
            if field == "Filed":
                html_val  = clean(str(html_row.get("FilingDate", "")) + " " + str(html_row.get("Jurisdiction", "")))
                excel_val = clean(str(excel_row.get("FilingDate", "")) + " " + str(excel_row.get("Jurisdiction", "")))
            else:
                html_val  = clean(str(html_row.get(field, "")))
                excel_val = clean(str(excel_row.get(field, "")))
            error_rows.append({
                "CaseNumber": case_number,
                "Field":      field,
                "HTML":       html_val,
                "Excel":      excel_val,
            })
    if not error_rows:
        return pd.DataFrame(columns=["CaseNumber", "Field", "HTML", "Excel"])
    return pd.DataFrame(error_rows, columns=["CaseNumber", "Field", "HTML", "Excel"])

# =========================
# CLIENT INPUT vs EXCEL MATCH SHEET
# =========================
def build_client_excel_match_df(client_df, excel_df):
    if client_df is None or client_df.empty:
        return pd.DataFrame(columns=["ClientInputCaseNumber", "ExcelCaseNumber", "MatchStatus", "SourceExcel"])
    col = "CaseNumber" if "CaseNumber" in client_df.columns else client_df.columns[0]
    excel_lookup = {}
    for _, erow in excel_df.iterrows():
        norm = normalize_case(str(erow.get("CaseNumber", "")))
        if norm and norm not in excel_lookup:
            excel_lookup[norm] = erow
    rows = []
    for val in client_df[col]:
        raw = clean(str(val))
        if not raw:
            continue
        norm = normalize_case(raw)
        erow = excel_lookup.get(norm)
        if erow is not None:
            rows.append({
                "ClientInputCaseNumber": raw,
                "ExcelCaseNumber":       clean(str(erow.get("CaseNumber", ""))),
                "MatchStatus":           "MATCH",
                "SourceExcel":           clean(str(erow.get("SourceExcel", ""))),
            })
        else:
            rows.append({
                "ClientInputCaseNumber": raw,
                "ExcelCaseNumber":       "",
                "MatchStatus":           "MISMATCH",
                "SourceExcel":           "",
            })
    return pd.DataFrame(rows, columns=["ClientInputCaseNumber", "ExcelCaseNumber", "MatchStatus", "SourceExcel"])

# =========================
# BUILD DATA FILL RATE COMPARISON SHEET
# =========================
def build_fill_rate_df(excel_df, active_fields):
    TARGET_RANGES = {
        "CaseNumber": (100, 100),
        "Citation": (90, 100),
        "FilingDate": (90, 100),
        "Jurisdiction": (90, 100),
        "CourtCaseType": (90, 100),
        "CaseStatus": (90, 100),
        "NameRaw": (90, 100),
        "StatuteDescription": (90, 100),
        "Pleadate": (25, 35),
        "Dispositiondate": (25, 35),
        "PleaNumber": (25, 35),
        "Plea": (25, 35),
        "DispositionNumber": (25, 35),
        "Disposition": (25, 35)
    }
    total_excel = len(excel_df)
    common_fields = [f for f in active_fields if f in excel_df.columns]
    rows = []
    for field in common_fields:
        excel_filled = excel_df[field].apply(
            lambda x: pd.notna(x) and str(x).strip() != ""
        ).sum()
        excel_pct = round((excel_filled / total_excel * 100), 2) if total_excel > 0 else 0
        t_min, t_max = TARGET_RANGES.get(field, (0, 10))
        if t_min == t_max:
            target_str = f"{t_min}%"
        else:
            target_str = f"{t_min}-{t_max}%"
        if excel_pct < t_min:
            status = "Below Range"
        elif excel_pct > t_max:
            status = "Above Range"
        else:
            status = "Within Range"
        rows.append({
            "Field": field,
            "Excel Fill Count": excel_filled,
            "Excel Fill %": excel_pct,
            "Expected Target Range": target_str,
            "Status": status,
        })
    if not rows:
        return pd.DataFrame(columns=["Field", "Excel Fill Count", "Excel Fill %", "Expected Target Range", "Status"])
    return pd.DataFrame(rows)

# =========================
# SAVE OUTPUT
# =========================
def save_output(output_file, html_df, excel_df, result_df, client_df, active_fields):
    errors_df = build_errors_df(result_df, html_df, excel_df)
    client_excel_df = build_client_excel_match_df(client_df, excel_df)
    fill_rate_df = build_fill_rate_df(excel_df, active_fields)
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        html_df.to_excel(writer,   sheet_name="HTML",   index=False)
        excel_df.to_excel(writer,  sheet_name="Excel",  index=False)
        result_df.to_excel(writer, sheet_name="Result", index=False)
        errors_df.to_excel(writer, sheet_name="Observation", index=False)
        fill_rate_df.to_excel(writer, sheet_name="FillRate_Comparison", index=False)
        if client_df is not None and not client_df.empty:
            client_df.to_excel(writer, sheet_name="ClientInput", index=False)
        if not client_excel_df.empty:
            client_excel_df.to_excel(writer, sheet_name="ClientInput_vs_Excel", index=False)

    wb = load_workbook(output_file)

    # -- Style FillRate_Comparison sheet --
    if "FillRate_Comparison" in wb.sheetnames:
        ws_fr = wb["FillRate_Comparison"]
        header_fill = PatternFill(start_color="1D3A6A", end_color="1D3A6A", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        thin_border = Border(
            left=Side(style='thin', color='D1D5DB'),
            right=Side(style='thin', color='D1D5DB'),
            top=Side(style='thin', color='D1D5DB'),
            bottom=Side(style='thin', color='D1D5DB')
        )
        status_col_idx = None
        pct_col_idx = None
        for idx, cell in enumerate(ws_fr[1], 1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            if cell.value == "Status":
                status_col_idx = idx
            elif cell.value == "Excel Fill %":
                pct_col_idx = idx
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        for row in ws_fr.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            if status_col_idx:
                status_cell = row[status_col_idx - 1]
                pct_cell = row[pct_col_idx - 1] if pct_col_idx else None
                if status_cell.value == "Within Range":
                    status_cell.fill = green_fill
                    if pct_cell: pct_cell.fill = green_fill
                elif status_cell.value == "Below Range":
                    status_cell.fill = red_fill
                    if pct_cell: pct_cell.fill = red_fill
                elif status_cell.value == "Above Range":
                    status_cell.fill = yellow_fill
                    if pct_cell: pct_cell.fill = yellow_fill
        for col in ws_fr.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 4, 30)
            ws_fr.column_dimensions[column_letter].width = adjusted_width
        ws_fr.freeze_panes = "A2"
        ws_fr.auto_filter.ref = ws_fr.dimensions

    # -- Style all other sheets --
    for sheet_name in wb.sheetnames:
        if sheet_name == "FillRate_Comparison": continue
        ws = wb[sheet_name]
        if ws.max_row is None or ws.max_row < 1 or ws.max_column is None or ws.max_column < 1: continue
        for col in ws.columns:
            max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)
        for cell in ws[1]: cell.font = Font(bold=True)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        if ws.max_row >= 1 and ws.max_column >= 1:
            table_range = f"A1:{ws.cell(row=1, column=ws.max_column).column_letter}{ws.max_row}"
            safe_name = re.sub(r"[^A-Za-z0-9_]", "_", sheet_name) + "Table"
            if not safe_name[0].isalpha() and safe_name[0] != "_": safe_name = "_" + safe_name
            tbl = Table(displayName=safe_name, ref=table_range)
            tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
            ws.add_table(tbl)

    # Colour Result sheet
    ws = wb["Result"]
    green_fill  = PatternFill(start_color="C6EFCE", fill_type="solid")
    red_fill    = PatternFill(start_color="FFC7CE", fill_type="solid")
    orange_fill = PatternFill(start_color="FFD700", fill_type="solid")
    grey_fill   = PatternFill(start_color="E5E7EB", fill_type="solid")
    client_col_idx = status_col_idx = client_case_res_idx = None
    for cell in ws[1]:
        if cell.value == "ClientMatch": client_col_idx = cell.column
        if cell.value == "ValidationStatus": status_col_idx = cell.column
        if cell.value == "ClientCaseNumber_Result": client_case_res_idx = cell.column
    for row in ws.iter_rows(min_row=2):
        status_val = row[status_col_idx - 1].value if status_col_idx else ""
        for cell in row:
            if cell.value == "MATCH": cell.fill = green_fill
            elif cell.value == "MISMATCH": cell.fill = red_fill
            elif status_val == "No cases matched your search criteria": cell.fill = orange_fill
            elif cell.column == client_col_idx:
                if cell.value == "FOUND": cell.fill = green_fill
                elif cell.value == "NOT IN CLIENT": cell.fill = grey_fill
            if client_case_res_idx and cell.column == client_case_res_idx:
                if cell.value == "MATCH": cell.fill = green_fill
                elif cell.value == "MISMATCH": cell.fill = red_fill

    # -- Colour Errors sheet --
    if "Observation" in wb.sheetnames:
        ws_err = wb["Observation"]
        html_col_idx = excel_col_idx = None
        for cell in ws_err[1]:
            if cell.value == "HTML": html_col_idx = cell.column
            elif cell.value == "Excel": excel_col_idx = cell.column
        for row in ws_err.iter_rows(min_row=2):
            for cell in row:
                if cell.column == html_col_idx and cell.value not in (None, ""):
                    cell.fill = PatternFill(start_color="DDEBF7", fill_type="solid")
                elif cell.column == excel_col_idx and cell.value not in (None, ""):
                    cell.fill = PatternFill(start_color="FFC7CE", fill_type="solid")

    # -- Colour ClientInput_vs_Excel sheet --
    if "ClientInput_vs_Excel" in wb.sheetnames:
        ws_ce = wb["ClientInput_vs_Excel"]
        match_col_idx = None
        for cell in ws_ce[1]:
            if cell.value == "MatchStatus": match_col_idx = cell.column
        for row in ws_ce.iter_rows(min_row=2):
            status_val = row[match_col_idx - 1].value if match_col_idx else ""
            row_fill = green_fill if status_val == "MATCH" else red_fill
            for cell in row: cell.fill = row_fill
    wb.save(output_file)
